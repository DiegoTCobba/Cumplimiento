import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

st.set_page_config(
    page_title="Filtro de Clientes > 30K",
    layout="wide"
)

st.title("📊 Filtro de Clientes con Montos Mayores a 30K")
st.write("Carga uno o varios archivos Excel para identificar clientes con montos elevados.")

# Subida de archivos
uploaded_files = st.file_uploader(
    "📂 Sube uno o más archivos Excel",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

if uploaded_files:
    dataframes = []

    for file in uploaded_files:
        try:
            df = pd.read_excel(file)

            # Columnas por posición (B, C, D, I, M)
            columnas_interes = df.iloc[:, [1, 2, 3, 8, 12]].copy()
            columnas_interes.columns = [
                "DOCUMENTO",
                "NUMERO DE DOCUMENTO ",
                "NOMBRE",
                "REFERENCIA",
                "MONTO"
            ]

            # Tipos correctos
            columnas_interes["MONTO"] = pd.to_numeric(
                columnas_interes["MONTO"], errors="coerce"
            )

            # 🔐 CLAVE: REFERENCIA como TEXTO
            columnas_interes["REFERENCIA"] = columnas_interes["REFERENCIA"].astype(str)

            # Filtrar montos mayores a 30k
            filtrado = columnas_interes[columnas_interes["MONTO"] > 30000]

            # Trazabilidad
            filtrado["ARCHIVO DE ORIGEN"] = file.name

            dataframes.append(filtrado)

        except Exception as e:
            st.error(f"Error procesando el archivo {file.name}: {e}")

    if dataframes:
        resultado_final = pd.concat(dataframes, ignore_index=True)

        st.success(f"✅ Se encontraron {len(resultado_final)} clientes con montos > 30K")
        st.dataframe(resultado_final, use_container_width=True)

        # ===============================
        # Exportar Excel formateado
        # ===============================
        buffer = BytesIO()
        resultado_final.to_excel(buffer, index=False, engine="openpyxl")
        buffer.seek(0)
        
        wb = load_workbook(buffer)
        ws = wb.active
        
        # 1️⃣ Forzar REFERENCIA como texto (columna D)
        for col in ws.iter_cols(min_col=4, max_col=4, min_row=2):
            for cell in col:
                cell.number_format = "@"
        
        # 2️⃣ Ajustar ancho de columnas automáticamente
        from openpyxl.utils import get_column_letter
        
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
        
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
            ws.column_dimensions[column_letter].width = max_length + 3
        
        buffer_final = BytesIO()
        wb.save(buffer_final)
        buffer_final.seek(0)
        
        st.download_button(
            label="⬇️ Descargar resultado en Excel",
            data=buffer_final,
            file_name="clientes_mayores_30k.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:
        st.warning("No se encontraron registros con montos mayores a 30K.")

